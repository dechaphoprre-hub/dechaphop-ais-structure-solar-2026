import openpyxl
import os
import json
import sys
import urllib.request
import csv
from datetime import datetime

# Reconfigure stdout to support UTF-8 on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQm0CGw3poRvXcSVv33aSs1QbSqA9OmJXIIFv6JCAcJC-ZE-iLP1TRugoZCQFNgn1ygeiRGfpC04rLQ/pub?gid=2088776154&single=true&output=csv"

def fetch_from_google_sheets():
    print("Fetching data from Google Sheets CSV...")
    try:
        req = urllib.request.Request(
            CSV_URL, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        )
        with urllib.request.urlopen(req, timeout=10) as response:
            csv_data = response.read().decode('utf-8')
        
        reader = csv.DictReader(csv_data.splitlines())
        drawings = []
        for row in reader:
            cleaned_row = {}
            for k, v in row.items():
                if k:
                    cleaned_row[k.strip()] = v.strip() if v else ""
            
            val = cleaned_row.get('Material Code')
            if val:
                try:
                    f_val = float(val)
                    if f_val.is_integer():
                        cleaned_row['Material Code'] = str(int(f_val))
                except ValueError:
                    pass
            
            if cleaned_row.get('Structure Code') or cleaned_row.get('Google Drive File URL'):
                drawings.append(cleaned_row)
                
        print(f"Successfully fetched {len(drawings)} drawing entries from Google Sheets.")
        return drawings
    except Exception as e:
        print(f"Warning: Failed to fetch from Google Sheets: {e}")
        return None

def generate():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    js_output_path = os.path.join(script_dir, "data.js")
    
    # Try fetching from Google Sheets first
    drawings = fetch_from_google_sheets()
    
    # If Google Sheets fetch fails, fallback to local Excel file
    if drawings is None:
        excel_path = os.path.join(script_dir, "Hia_Solar_Drawing_Portal_Automation_v2.xlsx")
        if not os.path.exists(excel_path):
            print(f"Error: Excel fallback file not found at {excel_path}")
            return False
            
        print(f"Falling back to local Excel file: {excel_path}")
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            sheet_name = "Drawing List Database"
            if sheet_name not in wb.sheetnames:
                print(f"Error: Sheet '{sheet_name}' not found in Excel!")
                return False
                
            ws = wb[sheet_name]
            
            # Read headers from first row
            headers = [cell.value for cell in ws[1]]
            print("Detected headers:", [str(h) for h in headers])
            
            drawings = []
            for r in range(2, ws.max_row + 1):
                row_vals = [cell.value for cell in ws[r]]
                
                if not any(v is not None for v in row_vals):
                    continue
                    
                row_dict = {}
                for col_idx, h in enumerate(headers):
                    if h is not None:
                        val = row_vals[col_idx] if col_idx < len(row_vals) else None
                        if isinstance(val, str):
                            val = val.strip()
                        if h == 'Material Code' and val is not None:
                            if isinstance(val, float) and val.is_integer():
                                val = str(int(val))
                            else:
                                val = str(val)
                        row_dict[str(h)] = val
                
                if row_dict.get('Structure Code') or row_dict.get('Google Drive File URL'):
                    drawings.append(row_dict)
            print(f"Successfully loaded {len(drawings)} drawing entries from local Excel.")
        except Exception as e:
            print(f"Error reading local Excel fallback: {e}")
            return False
            
    # Prepare JS content
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    js_content = f"""// Auto-generated drawing list data
// Generated on: {timestamp}

const LAST_UPDATED = "{timestamp}";

const DRAWING_DATA = {json.dumps(drawings, ensure_ascii=False, indent=2)};
"""
    
    with open(js_output_path, 'w', encoding='utf-8') as f:
        f.write(js_content)
        
    print(f"Successfully generated database JavaScript: {js_output_path}")
    return True

if __name__ == "__main__":
    generate()
